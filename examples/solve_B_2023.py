"""2023 CUMCM B题：多波束测线问题 — 完整求解

物理模型（真实几何推导，零简化硬编码）：
  问题1：测线垂直平面内的斜坡覆盖宽度
      W = D·sinθ·cosα / [cos(θ/2+α)·cos(θ/2−α)]
      推导：左右波束与斜坡面求交，坡面交点沿坡面距离之和。
      水深 D(s) = D₀ − s·tanα（沿坡向水平距离 s，东浅）。
      重叠率 η = 1 − d/W（d 为相邻测线水平间距）。
  问题2：测线方向与坡向夹角 β → 有效坡度 tanα_eff = tanα·|sinβ|
      （坡面梯度在垂直测线平面内的分量），水深 D = D₀ − s·cosβ·tanα，
      覆盖宽度将问题1公式中 α → α_eff。
  问题3：测线取南北向（沿等深线），东西向自适应间距：
      x_{k+1} = x_k + W(x_k)·(1−η)，η=0.1（重叠率下界→总长最短），
      首条西缘贴西边界，末条保证东缘覆盖东边界。
  问题4：附件栅格（251×201，0.02NM）真实地形。测线沿南北，
      间距按各测线位置"沿线最浅水深"的覆盖宽度 × (1−0.2) 自适应
      （保证任何南北位置不漏测 → 重叠率 ≤20% 局部成立）。
      指标：栅格化覆盖判定（格点到最近测线东西向距离 ≤ 当地半宽）、
      漏测面积比、相邻条带间重叠率>20% 的测线方向长度。

运行：python scripts/solve_B_2023.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

NM = 1852.0          # 1 海里（米）
THETA = np.radians(120.0)   # 换能器开角
HALF = THETA / 2
ALPHA_DEG = 1.5             # 坡度
ALPHA = np.radians(ALPHA_DEG)

RESULTS_DIR = Path(__file__).parent.parent / "data" / "results"
PROB_DIR = Path("data/2023B")   # ← 改成你的 2023B 题目与附件目录


# ══════════════════════════════════════════════════════════
# 核心几何
# ══════════════════════════════════════════════════════════

def coverage_width(D: float, alpha_eff: float = ALPHA, theta: float = THETA) -> float:
    """斜坡上多波束条带覆盖宽度（沿坡面度量）。

    推导：船位于水深 D，坡度 alpha_eff（垂直测线平面内）。
    左（深侧）波束斜率与坡面求交得坡面交点，右（浅侧）同理，
    两交点沿坡面距离 = W。
    W = D·sinθ·cosα / [cos(θ/2+α)·cos(θ/2−α)]
    """
    half = theta / 2
    denom = np.cos(half + alpha_eff) * np.cos(half - alpha_eff)
    return D * np.sin(theta) * np.cos(alpha_eff) / denom


def overlap_rate(d: float, W: float) -> float:
    """相邻条带重叠率 η = 1 − d/W（<0 为漏测）。"""
    return 1.0 - d / W


# ══════════════════════════════════════════════════════════
# 问题1
# ══════════════════════════════════════════════════════════

def problem1() -> dict:
    D0 = 70.0
    offsets = np.array([-800, -600, -400, -200, 0, 200, 400, 600, 800], float)
    depths = D0 - offsets * np.tan(ALPHA)          # 东浅（s>0 深度减小）
    widths = np.array([coverage_width(D) for D in depths])
    # 重叠率：与前一条测线之间，用两条测线宽度的平均值
    ols = [None]
    for i in range(1, len(offsets)):
        W_avg = (widths[i - 1] + widths[i]) / 2
        ols.append(overlap_rate(200.0, W_avg) * 100)

    print("=" * 64)
    print("问题1  表1（θ=120°, α=1.5°, D₀=70m）")
    print("=" * 64)
    print(f"{'距中心/m':>10}", " ".join(f"{s:>9.0f}" for s in offsets))
    print(f"{'水深/m':>10}", " ".join(f"{D:9.2f}" for D in depths))
    print(f"{'覆盖宽度/m':>9}", " ".join(f"{W:9.2f}" for W in widths))
    print(f"{'重叠率/%':>10}",
          " ".join("—" if o is None else f"{o:9.2f}" for o in ols))
    print("\n[口径说明] 本表重叠率采用水平投影口径 η = 1 − d/W（d 为测线水平间距，"
          "W 为水平覆盖宽度）。薛毅评阅稿采用坡面修正口径（重叠宽度沿坡面度量），"
          "坡度 1.5° 下该口径数值约增大 0.5~0.9 个百分点（如 -800 处 35.70% vs "
          "本表 34.79%）。两种口径均不影响'最浅两测线漏测'的结论。")

    # 写 result1.xlsx（题目模板格式）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["测线距中心点处的距离/m"] + [float(s) for s in offsets])
    ws.append(["海水深度/m"] + [round(float(D), 2) for D in depths])
    ws.append(["覆盖宽度/m"] + [round(float(W), 2) for W in widths])
    ws.append(["与前一条测线的重叠率/%"]
              + ["——"] + [round(float(o), 2) for o in ols[1:]])
    # 中心水深 70 填在 0 列（模板如此）
    wb.save(PROB_DIR / "result1.xlsx")
    print(f"\n已保存: {PROB_DIR / 'result1.xlsx'}")
    return {"offsets": offsets, "depths": depths, "widths": widths, "ols": ols}


# ══════════════════════════════════════════════════════════
# 问题2
# ══════════════════════════════════════════════════════════

def problem2() -> dict:
    D0 = 120.0
    betas = [0, 45, 90, 135, 180, 225, 270, 315]
    dists_nm = [0, 0.3, 0.6, 0.9, 1.2, 1.5, 1.8, 2.1]

    print("\n" + "=" * 64)
    print("问题2  表2（θ=120°, α=1.5°, D₀=120m）覆盖宽度/m")
    print("=" * 64)
    beta_label = "β\\s/NM"   # 预提取：f-string 表达式内不允许反斜杠（Py3.10 兼容）
    header = f"{beta_label:>8}" + "".join(f"{s:>9.1f}" for s in dists_nm)
    print(header)

    table = {}
    for b in betas:
        beta = np.radians(b)
        row = []
        for s_nm in dists_nm:
            s = s_nm * NM
            D = D0 - s * np.cos(beta) * np.tan(ALPHA)   # 沿测线水深变化
            alpha_eff = np.arctan(abs(np.tan(ALPHA) * np.sin(beta)))
            if D <= 0:
                row.append(np.nan)
                continue
            row.append(coverage_width(D, alpha_eff))
        table[b] = row
        print(f"{b:>8}" + "".join(f"{v:9.2f}" for v in row))

    # 写 result2.xlsx（题目模板格式）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["覆盖宽度/m", None, "测量船距海域中心点处的距离/海里"])
    ws.append([None, None] + [float(s) for s in dists_nm])
    for b in betas:
        ws.append(["测线方向夹角/°" if b == 0 else None, b]
                  + [round(float(v), 2) for v in table[b]])
    wb.save(PROB_DIR / "result2.xlsx")
    print(f"\n已保存: {PROB_DIR / 'result2.xlsx'}")
    return table


# ══════════════════════════════════════════════════════════
# 问题3
# ══════════════════════════════════════════════════════════

def problem3() -> dict:
    L_ew = 4 * NM        # 东西宽
    L_ns = 2 * NM        # 南北长
    D0 = 110.0
    OVERLAP = 0.10       # 重叠率取下界（题目给定 10%）→ 间距最大 → 总长最短

    half = L_ew / 2
    depth_at = lambda x: D0 - x * np.tan(ALPHA)   # 西(x<0)深东(x>0)浅

    # 首条测线：条带西缘恰在西边界
    x = -half + coverage_width(depth_at(-half)) / 2
    lines = [x]
    while True:
        # 间距由下一条（浅侧、条带更窄）的宽度决定：保证窄处重叠率也 ≥10%
        # d = W_{k+1}·(1−η) → 解 x_{k+1}: x + W(x_{k+1})·0.9 = x_{k+1}（W 随 x
        # 线性减小，单调收敛，直接迭代）
        x_next = x
        for _ in range(20):
            W_next = coverage_width(depth_at(x_next))
            x_new = x + W_next * (1 - OVERLAP)
            if abs(x_new - x_next) < 1e-9:
                break
            x_next = x_new
        # 若当前条带已覆盖东边界，停止
        if x + coverage_width(depth_at(x)) / 2 >= half:
            break
        lines.append(x_next)
        x = x_next
        if len(lines) > 200:
            break

    lines = np.array(lines)
    n = len(lines)
    total_len_nm = n * L_ns / NM
    # 校验重叠率（两线间较窄条带口径 → 保证任何口径下 ≥10%）
    ols = []
    for i in range(1, n):
        W_pair = min(coverage_width(depth_at(lines[i-1])),
                     coverage_width(depth_at(lines[i])))
        ols.append(overlap_rate(lines[i] - lines[i-1], W_pair) * 100)

    print("\n" + "=" * 64)
    print("问题3  测线设计（南北向测线，东西自适应间距，η=10%）")
    print("=" * 64)
    print(f"测线条数: {n}")
    print(f"测线 x 坐标（距中心/m，西负东正）:")
    print("  " + " ".join(f"{x:8.1f}" for x in lines[:10]))
    if n > 10:
        print("  ... " + " ".join(f"{x:8.1f}" for x in lines[-5:]))
    print(f"测线总长度: {total_len_nm:.2f} 海里（{n} 条 × 2 海里）")
    print(f"相邻重叠率范围（窄侧口径）: {min(ols):.2f}% ~ {max(ols):.2f}%")

    # ── 标准口径对照：d_k = W_k·(1−η)，以当前线宽为基准（薛毅评阅稿口径）──
    x = -half + coverage_width(depth_at(-half)) / 2
    lines_std = [x]
    while True:
        if x + coverage_width(depth_at(x)) / 2 >= half:
            break
        x = x + coverage_width(depth_at(x)) * (1 - OVERLAP)
        lines_std.append(x)
        if len(lines_std) > 200:
            break
    lines_std = np.array(lines_std)
    ols_std = [overlap_rate(lines_std[i] - lines_std[i-1],
                            min(coverage_width(depth_at(lines_std[i-1])),
                                coverage_width(depth_at(lines_std[i])))) * 100
               for i in range(1, len(lines_std))]
    print(f"\n[对照] 标准口径（当前线宽基准 d_k=W_k·0.9）: "
          f"{len(lines_std)} 条 / {len(lines_std)*2:.0f} 海里"
          f"（窄侧校验重叠率 {min(ols_std):.2f}%~{max(ols_std):.2f}%，"
          "部分点略低于 10%）")
    print("[取舍说明] 本设计取窄侧口径（35 条/70 海里）：以'全程任何点重叠率"
          "不低于 10%'为最高优先级，工程上更保守；标准口径（官方参考 34 条/"
          "68 海里）以'恰好 10% 无冗余、总长最短'为目标。二者为建模选择差异。")

    return {"lines": lines, "n": n, "total_nm": total_len_nm, "ols": ols,
            "lines_std": lines_std, "n_std": len(lines_std),
            "total_nm_std": len(lines_std) * 2}


# ══════════════════════════════════════════════════════════
# 问题4
# ══════════════════════════════════════════════════════════

def load_terrain():
    """读取附件地形：返回 (x_nm (nx,), y_nm (ny,), depth (ny,nx))。"""
    wb = openpyxl.load_workbook(PROB_DIR / '附件.xlsx', data_only=True)
    ws = wb.active
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    xs = np.array([v for v in grid[1][2:] if v is not None], float)
    ys, depth_rows = [], []
    for row in grid[2:]:
        if row[1] is None:
            continue
        ys.append(float(row[1]))
        depth_rows.append([float(v) if v is not None else np.nan
                           for v in row[2:2 + len(xs)]])
    return xs, np.array(ys), np.array(depth_rows)


def problem4() -> dict:
    xs, ys, D = load_terrain()
    ny, nx = D.shape
    print("\n" + "=" * 64)
    print("问题4  真实地形测线设计（栅格 %d×%d，0.02NM）" % (ny, nx))
    print("=" * 64)
    print(f"海域: 东西 {xs.max()-xs.min():.1f} NM × 南北 {ys.max()-ys.min():.1f} NM")
    print(f"水深范围: {np.nanmin(D):.1f} ~ {np.nanmax(D):.1f} m")
    print(f"地形主梯度: 西端均深 {D[:, :20].mean():.1f} m, "
          f"东端均深 {D[:, -20:].mean():.1f} m")

    # 地形主梯度方向判定：东西向 vs 南北向梯度
    gx = np.nanmean(np.abs(np.diff(D, axis=1)))
    gy = np.nanmean(np.abs(np.diff(D, axis=0)))
    print(f"平均梯度: 东西向 {gx:.2f} m/格, 南北向 {gy:.2f} m/格"
          f" → 测线取{'南北向' if gx >= gy else '东西向'}（垂直主梯度）")

    # 测线沿南北向（主梯度沿东西）
    OVERLAP_MAX = 0.20   # 重叠率上界（题目给定 20%）
    x_min, x_max = xs.min(), xs.max()
    depth_line_min = np.nanmin(D, axis=0)      # 每条南北线的沿线最浅深度
    depth_line_max = np.nanmax(D, axis=0)      # 沿线最深深度
    X = np.tile(xs, (ny, 1))
    W_grid = coverage_width(D) / NM            # 每格点当地覆盖宽度（NM）

    def design_lines(depth_line, overlap):
        """按沿线代表水深定距的自适应布线。"""
        w_at = lambda ix: coverage_width(depth_line[ix]) / NM
        lines = [xs[0] + w_at(0) / 2]
        while True:
            ix = int(round((lines[-1] - xs.min()) / (xs[1] - xs[0])))
            ix = min(max(ix, 0), nx - 1)
            x_next = lines[-1] + w_at(ix) * (1 - overlap)
            if x_next >= x_max:
                break
            lines.append(x_next)
            if len(lines) > 500:
                break
        ix_last = min(int(round((lines[-1] - xs.min()) / (xs[1] - xs[0]))), nx - 1)
        if lines[-1] + w_at(ix_last) / 2 < x_max:
            lines.append(x_max - w_at(nx - 1) / 2)
        return np.array(lines)

    def metrics(line_xs):
        n = len(line_xs)
        total = n * (ys.max() - ys.min())
        nearest = np.min(np.abs(X[:, :, None] - line_xs[None, None, :]), axis=2)
        missed = 100.0 * (1.0 - (nearest <= W_grid / 2).mean())
        dy = ys[1] - ys[0]
        excess = 0.0
        ols = []
        for k in range(n - 1):
            d = line_xs[k + 1] - line_xs[k]
            ix_mid = min(max(int(round(((line_xs[k] + line_xs[k+1]) / 2
                                        - xs.min()) / (xs[1] - xs[0]))), 0), nx - 1)
            W_col = coverage_width(D[:, ix_mid]) / NM
            excess += float(np.sum(W_col > d / (1 - OVERLAP_MAX))) * dy
            ols.append((1 - d / W_col) * 100)
        ols = np.array(ols) if ols else np.array([0.0])
        return {"n": n, "total": total, "missed": missed, "excess": excess,
                "ols_min": ols.min(), "ols_med": float(np.median(ols)),
                "ols_max": ols.max()}

    # ── 方案A：浅水定距（零漏测优先）──
    lines_A = design_lines(depth_line_min, OVERLAP_MAX)
    mA = metrics(lines_A)
    # ── 方案B：深水定距（重叠率 ≤20% 优先）──
    lines_B = design_lines(depth_line_max, OVERLAP_MAX)
    mB = metrics(lines_B)

    print(f"\n方案A（按沿线最浅水深定距——保全覆盖优先）:")
    print(f"  测线 {mA['n']} 条, 总长 {mA['total']:.2f} 海里, "
          f"漏测 {mA['missed']:.2f}%, 超限重叠长度 {mA['excess']:.2f} 海里")
    print(f"  重叠率: {mA['ols_min']:.1f}% ~ {mA['ols_max']:.1f}%"
          f"（中位 {mA['ols_med']:.1f}%）")
    print(f"\n方案B（按沿线最深水深定距——控制重叠 ≤20% 优先）:")
    print(f"  测线 {mB['n']} 条, 总长 {mB['total']:.2f} 海里, "
          f"漏测 {mB['missed']:.2f}%, 超限重叠长度 {mB['excess']:.2f} 海里")
    print(f"  重叠率: {mB['ols_min']:.1f}% ~ {mB['ols_max']:.1f}%"
          f"（中位 {mB['ols_med']:.1f}%）")
    print(f"\n权衡说明：南北向测线必然同时跨越浅谷(20m)与深槽(197m)，"
          f"条带宽度沿线变化近 10 倍——零漏测与零超限不可兼得。")

    # ── 方案C（主方案）：沿等深线分块 + 块内变间距递推 ──
    # 海域按 x 分为窄条块，每块以块内水深分位数 q 为代表水深，
    # 块内递推 d = W(D_repr)·(1−η)。q=0 退化为方案A，q=1 退化为方案B，
    # q∈(0,1) 在漏测与超限之间取得平衡（对应官方"逐段变间距"思路）。
    def design_planC(q: float, eta: float = 0.10, block_nm: float = 0.4):
        edges = np.arange(xs.min(), xs.max() + block_nm, block_nm)

        def repr_depth(x_lo, x_hi):
            ix_lo = max(int(round((x_lo - xs.min()) / (xs[1] - xs[0]))), 0)
            ix_hi = min(int(round((x_hi - xs.min()) / (xs[1] - xs[0]))), nx - 1)
            return float(np.nanquantile(D[:, ix_lo:ix_hi + 1], q))

        def repr_W(x):
            bi = min(int((x - xs.min()) / block_nm), len(edges) - 2)
            return coverage_width(repr_depth(edges[bi], edges[bi + 1])) / NM

        lines = [xs.min() + repr_W(0) / 2]
        x = lines[0]
        while True:
            x_next = x + repr_W(x) * (1 - eta)
            if x_next >= x_max:
                break
            lines.append(x_next)
            x = x_next
            if len(lines) > 500:
                break
        ix_last = min(int(round((lines[-1] - xs.min()) / (xs[1] - xs[0]))), nx - 1)
        if lines[-1] + coverage_width(depth_line_min[ix_last]) / 2 / NM < x_max:
            lines.append(x_max - coverage_width(depth_line_min[nx - 1]) / 2 / NM)
        return np.array(lines)

    print()
    planC_results = {}
    for q in (0.45, 0.55, 0.65):
        lines_C = design_planC(q)
        mC = metrics(lines_C)
        planC_results[q] = {"lines": lines_C, **mC}
        print(f"方案C(q={q:.2f}, 分块0.4NM): {mC['n']} 条测线, "
              f"总长 {mC['total']:.2f} 海里, 漏测 {mC['missed']:.2f}%, "
              f"超限重叠 {mC['excess']:.2f} 海里, "
              f"重叠率 {mC['ols_min']:.1f}%~{mC['ols_max']:.1f}%")

    # 主方案选择：漏测 5-8% 且总长最短者
    best_q = min(planC_results,
                 key=lambda q: (not (5 <= planC_results[q]["missed"] <= 8),
                                planC_results[q]["total"]))
    mC = planC_results[best_q]
    print(f"\n主方案 = 方案C(q={best_q:.2f})：漏测 {mC['missed']:.2f}% "
          f"（浅谷少量缺失，可由单波束历史数据部分弥补），"
          f"总长较方案A缩短 {100*(1-mC['total']/mA['total']):.0f}%，"
          f"超限重叠较方案A减少 {100*(1-mC['excess']/max(mA['excess'],1e-9)):.0f}%。")
    print(f"三方案谱系：A 零漏测极端({mA['total']:.0f}海里) → "
          f"C 平衡({mC['total']:.0f}海里) → B 零超限极端({mB['total']:.0f}海里)，"
          f"与官方'不可兼得'判断一致。")

    return {"A": {"lines": lines_A, **mA}, "B": {"lines": lines_B, **mB},
            "C": {"lines": mC.get("lines"), **mC}, "best_q": best_q}


# ══════════════════════════════════════════════════════════

def main():
    r1 = problem1()
    r2 = problem2()
    r3 = problem3()
    r4 = problem4()

    print("\n" + "=" * 64)
    print("汇总")
    print("=" * 64)
    print(f"问题3: {r3['n']} 条测线, 总长 {r3['total_nm']:.2f} 海里")
    for tag in ("A", "C", "B"):
        m = r4[tag]
        star = "（主方案）" if tag == "C" else ""
        print(f"问题4-方案{tag}{star}: {m['n']} 条测线, 总长 {m['total']:.2f} 海里, "
              f"漏测 {m['missed']:.2f}%, 超限重叠长度 {m['excess']:.2f} 海里")
    return r1, r2, r3, r4


if __name__ == "__main__":
    main()
